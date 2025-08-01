import os
from openpyxl import load_workbook

def read_excel_data():
    file_path = os.path.join(os.path.abspath(os.curdir), "testdata", "Opencart_LoginData.xlsx")
    workbook = load_workbook(file_path)
    sheet = workbook["Sheet1"]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        username, password, expected_result = row
        data.append((username, password, expected_result))
    return data